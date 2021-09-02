import pyqrcode
import os
from os import path
from openpyxl import load_workbook
import glob

# Get user input
# Note: program assumes first row is just column titles
FILENAME = input('Name of excel file: ')
WB_TITLE = input('Workbook title: ')
QR_COL = int(input('Column # to make QRs: '))

# Load the excel sheet
wb = load_workbook(filename = FILENAME)
ws = wb[WB_TITLE]

# Make the QR directory if it doesn't exist
if(path.exists('QRs')):
    files = glob.glob('QRs/*.png')
    for f in files:
        os.remove(f)
else:
    os.mkdir('QRs')

# Read through the specified column and create the unique QRs
for col in ws.iter_cols(min_row=2, min_col=QR_COL, max_col=QR_COL):
    for cell in col:
        qr = pyqrcode.create(cell.value)
        with open('QRs/' + cell.value + '.png', 'wb') as fstream:
            qr.png(fstream, scale=5)

wb.close()